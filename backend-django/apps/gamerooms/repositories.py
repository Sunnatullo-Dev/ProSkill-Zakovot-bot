"""Online O'yin Xonasi — barcha biznes logika.

Hammasi atomic (transaction.atomic + select_for_update kerakli joylarda).
Race condition'lardan himoyalangan:
  - Deadline o'tganda ham kelgan javob → reject (server vaqti authoritative)
  - Savol yopilayotganda va javob kelayotganda → atomic gate
  - Ikki marta baholash → idempotent + atomic update
  - Xona tugaganida scoring → try_finish + select_for_update

Leak oldini olish:
  - Aktiv savol payloadi ishtirokchiga boshqalarning javob matni ko'rsatilmaydi
  - To'g'ri javob FAQAT savol closed bo'lganida yoki admin so'raganda ochiladi
"""
from __future__ import annotations

import logging
import secrets
import time
from typing import Any

from django.contrib.auth.hashers import check_password, identify_hasher, make_password
from django.db import IntegrityError, transaction
from django.db.models import F as models_F
from django.utils import timezone

from apps.core.exceptions import AppError

from .models import (
    GameQuestion,
    GameRoom,
    Participant,
    Submission,
)


logger = logging.getLogger(__name__)


# ─── Konstantalar ─────────────────────────────────────────────────────────────

GRACE_MS = 2_000                  # Deadline'dan keyin qabul qilingan javob uchun
CODE_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"  # I/O/0/1 chiqarilgan
CODE_LENGTH = 6
CODE_GEN_ATTEMPTS = 16

VALID_TIME_LIMITS = {30, 60, 90, 120, 180}
VALID_POINT_VALUES = {1, 2, 3}
VALID_QUESTION_TYPES = {"text", "audio", "image"}

# leaderboard'dagi top N g'oliblar
TOP_WINNERS_COUNT = 3


def _now_ms() -> int:
    return int(time.time() * 1000)


def _generate_code() -> str:
    return "".join(secrets.choice(CODE_ALPHABET) for _ in range(CODE_LENGTH))


def _hash_room_password(raw: str) -> str:
    raw = (raw or "").strip()
    return make_password(raw) if raw else ""


def _is_hashed_password(value: str) -> bool:
    if not value:
        return False
    try:
        identify_hasher(value)
        return True
    except Exception:  # noqa: BLE001
        return False


def _room_password_matches(room: GameRoom, candidate: str) -> bool:
    stored = room.join_password or ""
    candidate = (candidate or "").strip()
    if not stored:
        return True
    if _is_hashed_password(stored):
        return check_password(candidate, stored)
    if candidate == stored:
        room.join_password = _hash_room_password(candidate)
        room.save(update_fields=["join_password"])
        return True
    return False


# ─── Xona boshqaruvi ─────────────────────────────────────────────────────────

def create_room(
    *,
    admin_telegram_id: int,
    name: str,
    join_password: str = "",
    extra_admin_ids: list[dict] | None = None,
) -> dict[str, Any]:
    """Yangi xona yaratadi. Adminni qaytaradi."""
    name = (name or "").strip()
    if not name:
        raise AppError(400, "Xona nomi bo'sh bo'lmasligi kerak")
    if len(name) > 200:
        raise AppError(400, "Xona nomi 200 ta belgidan oshmasligi kerak")
    if join_password and len(join_password) > 50:
        raise AppError(400, "Parol 50 ta belgidan oshmasligi kerak")

    code = ""
    for _ in range(CODE_GEN_ATTEMPTS):
        candidate = _generate_code()
        if not GameRoom.objects.filter(code=candidate).exists():
            code = candidate
            break
    if not code:
        raise AppError(500, "Xona kodi yaratib bo'lmadi, qaytadan urinib ko'ring")

    room = GameRoom.objects.create(
        code=code,
        name=name,
        admin_telegram_id=admin_telegram_id,
        join_password=_hash_room_password(join_password),
        extra_admin_ids=extra_admin_ids or [],
        status="waiting",
    )
    return {
        "code": room.code,
        "name": room.name,
        "status": room.status,
        "adminTelegramId": room.admin_telegram_id,
        "hasPassword": bool(room.join_password),
        "createdAt": room.created_at.isoformat(),
        "startedAt": None,
        "finishedAt": None,
        "participantCount": 0,
        "leaderboard": [],
        "currentQuestion": None,
        "viewerTelegramId": admin_telegram_id,
        "viewerIsAdmin": True,
    }


def get_room(code: str) -> GameRoom:
    room = GameRoom.objects.filter(code=code.strip().upper()).first()
    if not room:
        raise AppError(404, "Xona topilmadi")
    return room


def join_room(
    *,
    code: str,
    telegram_id: int,
    display_name: str,
    join_password: str = "",
) -> dict[str, Any]:
    """Ishtirokchi xonaga qo'shiladi yoki qaytib ulanadi."""
    if telegram_id <= 0:
        raise AppError(401, "Faqat Telegram foydalanuvchilar qo'shilishi mumkin")

    display_name = (display_name or "").strip()
    if not display_name:
        raise AppError(400, "Taxallus (display_name) bo'sh bo'lmasligi kerak")
    if len(display_name) > 120:
        raise AppError(400, "Taxallus 120 ta belgidan oshmasligi kerak")

    normalized = code.strip().upper()
    with transaction.atomic():
        room = (
            GameRoom.objects.select_for_update()
            .filter(code=normalized)
            .first()
        )
        if not room:
            raise AppError(404, "Bu kod bilan xona topilmadi")
        if room.status == "finished":
            raise AppError(409, "Xona tugagan — kirib bo'lmaydi")

        # Parol tekshirish
        if room.join_password:
            if not _room_password_matches(room, join_password):
                raise AppError(403, "Noto'g'ri parol")

        existing = Participant.objects.filter(
            room=room, telegram_id=telegram_id
        ).first()
        if existing:
            # Qaytib ulanish — taxallus yangilanishi mumkin
            if display_name != existing.display_name:
                existing.display_name = display_name
                existing.save(update_fields=["display_name", "last_seen_at"])
        else:
            try:
                Participant.objects.create(
                    room=room,
                    telegram_id=telegram_id,
                    display_name=display_name,
                )
            except IntegrityError:
                # Race: boshqa worker shu telegram_id ni qo'shib qo'ydi
                raise AppError(409, "Allaqachon qo'shilgansiz")

    return get_room_state(normalized, viewer_telegram_id=telegram_id)


def start_room(*, code: str, admin_telegram_id: int) -> dict[str, Any]:
    """Admin xonani 'waiting' → 'active' qiladi."""
    normalized = code.strip().upper()
    with transaction.atomic():
        room = GameRoom.objects.select_for_update().filter(code=normalized).first()
        if not room:
            raise AppError(404, "Xona topilmadi")
        _require_room_admin(room, admin_telegram_id)
        if room.status != "waiting":
            raise AppError(409, f"Xona '{room.status}' holatda — boshlash mumkin emas")
        if room.participants.count() == 0:
            raise AppError(409, "Kamida 1 ta ishtirokchi kerak")
        room.status = "active"
        room.started_at = timezone.now()
        room.save(update_fields=["status", "started_at"])
    return get_room_state(normalized, viewer_telegram_id=admin_telegram_id)


def finish_room(*, code: str, admin_telegram_id: int) -> dict[str, Any]:
    """Admin xonani yakunlaydi. Joriy aktiv savol yopiladi."""
    normalized = code.strip().upper()
    with transaction.atomic():
        room = GameRoom.objects.select_for_update().filter(code=normalized).first()
        if not room:
            raise AppError(404, "Xona topilmadi")
        _require_room_admin(room, admin_telegram_id)
        if room.status == "finished":
            return get_room_state(normalized, viewer_telegram_id=admin_telegram_id)
        # Aktiv savol bo'lsa yopamiz
        if room.current_question_id:
            GameQuestion.objects.filter(
                id=room.current_question_id, status="active"
            ).update(status="closed", closed_at=timezone.now())
        room.status = "finished"
        room.finished_at = timezone.now()
        room.save(update_fields=["status", "finished_at"])
    return get_room_state(normalized, viewer_telegram_id=admin_telegram_id)


# ─── Savol boshqaruvi ─────────────────────────────────────────────────────────

def push_question(
    *,
    code: str,
    admin_telegram_id: int,
    question_type: str,
    body: str,
    media_ref: str = "",
    caption: str = "",
    correct_answer: str = "",
    time_limit_seconds: int = 120,
    point_value: int = 1,
    is_bonus: bool = False,
    is_quick: bool = False,
) -> dict[str, Any]:
    """Admin yangi savol qo'shadi va darhol 'active' qiladi.

    Oldingi aktiv savol bo'lsa — avval uni yopish kerak (close_question).
    """
    normalized = code.strip().upper()

    # Validatsiyalar (DB ga kirishdan oldin)
    question_type = (question_type or "text").strip()
    if question_type not in VALID_QUESTION_TYPES:
        raise AppError(400, f"Savol turi noto'g'ri: text/audio/image bo'lishi kerak")
    body = (body or "").strip()
    if not body:
        raise AppError(400, "Savol matni bo'sh bo'lmasligi kerak")
    if len(body) > 2000:
        raise AppError(400, "Savol matni 2000 ta belgidan oshmasligi kerak")
    if time_limit_seconds not in VALID_TIME_LIMITS:
        raise AppError(400, f"Vaqt limiti 30/60/90/120/180 dan biri bo'lishi kerak")
    if point_value not in VALID_POINT_VALUES:
        raise AppError(400, f"Ball qiymati 1/2/3 dan biri bo'lishi kerak")

    with transaction.atomic():
        room = GameRoom.objects.select_for_update().filter(code=normalized).first()
        if not room:
            raise AppError(404, "Xona topilmadi")
        _require_room_admin(room, admin_telegram_id)
        if room.status != "active":
            raise AppError(409, f"Xona '{room.status}' holatda — savol qo'shing mumkin emas")

        # Oldingi aktiv savol bo'lsa — rad etamiz
        active_q = GameQuestion.objects.filter(room=room, status="active").first()
        if active_q:
            raise AppError(
                409,
                "Oldingi savol hali ochiq. Avval uni yoping (close_question)."
            )

        # Tartib raqami
        last = (
            GameQuestion.objects.filter(room=room)
            .order_by("-order_index")
            .values_list("order_index", flat=True)
            .first()
        )
        order_index = (last or 0) + 1

        now = timezone.now()
        question = GameQuestion.objects.create(
            room=room,
            question_type=question_type,
            body=body,
            media_ref=(media_ref or "").strip()[:500],
            caption=(caption or "").strip()[:500],
            correct_answer=(correct_answer or "").strip(),
            time_limit_seconds=time_limit_seconds,
            point_value=point_value,
            order_index=order_index,
            is_bonus=bool(is_bonus),
            is_quick=bool(is_quick),
            status="active",
            activated_at=now,
        )

        room.current_question = question
        room.save(update_fields=["current_question"])

    return _serialize_question_for_admin(question)


def close_question(
    *,
    code: str,
    admin_telegram_id: int,
    question_id: int,
) -> dict[str, Any]:
    """Admin aktiv savolni yopadi. Javoblar qabul qilinmay qoladi."""
    normalized = code.strip().upper()
    with transaction.atomic():
        room = GameRoom.objects.select_for_update().filter(code=normalized).first()
        if not room:
            raise AppError(404, "Xona topilmadi")
        _require_room_admin(room, admin_telegram_id)

        question = GameQuestion.objects.select_for_update().filter(
            id=question_id, room=room
        ).first()
        if not question:
            raise AppError(404, "Savol topilmadi")
        if question.status == "closed":
            return _serialize_question_for_admin(question)
        if question.status != "active":
            raise AppError(409, f"Savol '{question.status}' holatda — yopib bo'lmaydi")

        question.status = "closed"
        question.closed_at = timezone.now()
        question.save(update_fields=["status", "closed_at"])

        # Agar bu joriy savol bo'lsa — current_question ni null qilamiz
        if room.current_question_id == question.id:
            room.current_question = None
            room.save(update_fields=["current_question"])

    return _serialize_question_for_admin(question)


def set_correct_answer(
    *,
    code: str,
    admin_telegram_id: int,
    question_id: int,
    correct_answer: str,
) -> dict[str, Any]:
    """Admin to'g'ri javobni belgilaydi yoki o'zgartiradi."""
    normalized = code.strip().upper()
    correct_answer = (correct_answer or "").strip()
    if not correct_answer:
        raise AppError(400, "To'g'ri javob bo'sh bo'lmasligi kerak")

    with transaction.atomic():
        room = GameRoom.objects.select_for_update().filter(code=normalized).first()
        if not room:
            raise AppError(404, "Xona topilmadi")
        _require_room_admin(room, admin_telegram_id)

        question = GameQuestion.objects.select_for_update().filter(
            id=question_id, room=room
        ).first()
        if not question:
            raise AppError(404, "Savol topilmadi")

        question.correct_answer = correct_answer
        question.save(update_fields=["correct_answer"])

    return _serialize_question_for_admin(question)


# ─── Javob berish ─────────────────────────────────────────────────────────────

def submit_answer(
    *,
    code: str,
    telegram_id: int,
    question_id: int,
    answer_text: str,
) -> dict[str, Any]:
    """Ishtirokchi savol uchun javob yuboradi yoki tahrirlaydi (upsert).

    Bitta ishtirokchi bitta savol uchun faqat bitta submission'ga ega.
    Deadline'gacha tahrirlash mumkin — har safar `updated_at` yangilanadi.
    """
    answer_text = (answer_text or "").strip()
    if not answer_text:
        raise AppError(400, "Javob bo'sh bo'lmasligi kerak")
    if len(answer_text) > 500:
        raise AppError(400, "Javob 500 ta belgidan oshmasligi kerak")

    normalized = code.strip().upper()
    now_ms = _now_ms()

    with transaction.atomic():
        room = GameRoom.objects.select_for_update().filter(code=normalized).first()
        if not room:
            raise AppError(404, "Xona topilmadi")
        if room.status != "active":
            raise AppError(409, "Xona aktiv emas — javob qabul qilinmaydi")

        question = GameQuestion.objects.select_for_update().filter(
            id=question_id, room=room
        ).first()
        if not question:
            raise AppError(404, "Savol topilmadi")
        if question.status != "active":
            raise AppError(409, "Bu savol aktiv emas — javob qabul qilinmaydi")

        # Server-authoritative deadline tekshirish + GRACE_MS
        if question.activated_at:
            deadline_ms = (
                int(question.activated_at.timestamp() * 1000)
                + question.time_limit_seconds * 1000
            )
            if now_ms > deadline_ms + GRACE_MS:
                raise AppError(409, "Vaqt tugadi — javob qabul qilinmaydi")

        participant = Participant.objects.select_for_update().filter(
            room=room, telegram_id=telegram_id
        ).first()
        if not participant:
            raise AppError(403, "Siz bu xonada ro'yxatdan o'tmagansiz")

        # Upsert: mavjud bo'lsa yangilash, yo'q bo'lsa yaratish
        existing = Submission.objects.select_for_update().filter(
            question=question, participant=participant
        ).first()
        if existing:
            existing.answer_text = answer_text
            # Qayta baholash kerak bo'lsa — is_correct va points reset
            if existing.graded_by:
                existing.is_correct = None
                existing.points_awarded = None
                existing.graded_by = ""
                existing.grading_note = ""
            existing.save(update_fields=[
                "answer_text", "is_correct", "points_awarded",
                "graded_by", "grading_note", "updated_at",
            ])
            sub = existing
        else:
            try:
                sub = Submission.objects.create(
                    question=question,
                    participant=participant,
                    answer_text=answer_text,
                )
            except IntegrityError:
                # Race: boshqa worker shu (question, participant) ni yaratib qo'ydi
                raise AppError(409, "Javob allaqachon yuborilgan — tahrirlash uchun qaytadan urinib ko'ring")

    return {
        "submissionId": sub.id,
        "questionId": question_id,
        "answerText": sub.answer_text,
        "submittedAt": sub.submitted_at.isoformat(),
        "updatedAt": sub.updated_at.isoformat(),
        "graded": sub.graded_by != "",
    }


# ─── Baholash ────────────────────────────────────────────────────────────────

def auto_grade_question(
    *,
    code: str,
    admin_telegram_id: int,
    question_id: int,
) -> dict[str, Any]:
    """Avtomatik baholash: exact_match + Gemini AI.

    `correct_answer` bo'sh bo'lsa — xato.
    Har bir submission atomik tarzda baholanadi va ishtirokchi bali yangilanadi.
    Idempotent: qayta chaqirilsa allaqachon baholangan submission'lar o'tkazib yuboriladi.
    """
    normalized = code.strip().upper()
    room = get_room(normalized)
    _require_room_admin(room, admin_telegram_id)

    question = GameQuestion.objects.filter(id=question_id, room=room).first()
    if not question:
        raise AppError(404, "Savol topilmadi")
    if question.status not in ("active", "closed"):
        raise AppError(409, "Savol hali push qilinmagan yoki baholash uchun tayyor emas")
    if not question.correct_answer:
        raise AppError(400, "To'g'ri javob kiritilmagan — avval set_correct_answer chaqiring")

    submissions = list(
        Submission.objects.select_related("participant")
        .filter(question=question)
    )
    if not submissions:
        return {
            "questionId": question_id,
            "gradedCount": 0,
            "skippedCount": 0,
            "correctCount": 0,
        }

    from apps.answers.grading import exact_match_grade
    from apps.answers.gemini import check_answer as gemini_check

    graded = 0
    skipped = 0
    correct = 0

    for sub in submissions:
        # Idempotent — allaqachon baholangan o'tkaziladi
        if sub.graded_by:
            skipped += 1
            continue

        # 1-qadam: exact match (tez, bepul)
        gr = exact_match_grade(sub.answer_text, question.correct_answer)
        if gr.status == "correct":
            is_correct = True
            note = "exact_match"
        else:
            # 2-qadam: Gemini AI fuzzy
            result = gemini_check(question.body, question.correct_answer, sub.answer_text)
            is_correct = result.status == "correct"
            note = result.explanation[:300] if result.explanation else result.status

        points = question.point_value if is_correct else 0

        with transaction.atomic():
            # Atomik gate: birinchi yozuvchi g'alaba qozonadi (race-safe)
            updated = Submission.objects.filter(
                id=sub.id, graded_by=""
            ).update(
                is_correct=is_correct,
                points_awarded=points,
                graded_by="auto",
                grading_note=note,
            )
            if updated:
                graded += 1
                if is_correct:
                    correct += 1
                    # Ishtirokchi balini va speed score'ni yangilash
                    _add_participant_points(
                        participant_id=sub.participant_id,
                        points=points,
                        submitted_at_ms=int(sub.submitted_at.timestamp() * 1000),
                        activated_at_ms=int(question.activated_at.timestamp() * 1000)
                        if question.activated_at
                        else 0,
                    )
            else:
                skipped += 1

    return {
        "questionId": question_id,
        "gradedCount": graded,
        "skippedCount": skipped,
        "correctCount": correct,
    }


def manual_grade_submission(
    *,
    code: str,
    admin_telegram_id: int,
    submission_id: int,
    is_correct: bool,
) -> dict[str, Any]:
    """Admin qo'lda bitta submission'ni to'g'ri/noto'g'ri belgilaydi.

    Qayta belgilash mumkin (override): ball farqi qo'lda hisoblanadi.
    """
    normalized = code.strip().upper()
    room = get_room(normalized)
    _require_room_admin(room, admin_telegram_id)

    with transaction.atomic():
        sub = Submission.objects.select_for_update().select_related(
            "question", "participant"
        ).filter(id=submission_id, question__room=room).first()
        if not sub:
            raise AppError(404, "Submission topilmadi")

        question = sub.question
        if question.room_id != room.id:
            raise AppError(403, "Bu submission boshqa xonaga tegishli")

        new_points = question.point_value if is_correct else 0
        old_points = sub.points_awarded or 0
        delta = new_points - old_points

        sub.is_correct = is_correct
        sub.points_awarded = new_points
        sub.graded_by = "manual"
        sub.grading_note = ""
        sub.save(update_fields=["is_correct", "points_awarded", "graded_by", "grading_note", "updated_at"])

        if delta != 0:
            # Participant ball farqini atomik yangilaymiz
            Participant.objects.filter(id=sub.participant_id).update(
                total_points=models_F("total_points") + delta
            )
            # Speed score yangilash (faqat yangi to'g'ri javob uchun)
            if is_correct and old_points == 0 and question.activated_at:
                activated_ms = int(question.activated_at.timestamp() * 1000)
                submitted_ms = int(sub.submitted_at.timestamp() * 1000)
                response_ms = max(0, submitted_ms - activated_ms)
                Participant.objects.filter(id=sub.participant_id).update(
                    speed_score_ms=models_F("speed_score_ms") + response_ms
                )

    return {
        "submissionId": sub.id,
        "isCorrect": is_correct,
        "pointsAwarded": new_points,
        "gradedBy": "manual",
    }


# ─── O'yin holati (polling endpoint) ─────────────────────────────────────────

def get_room_state(
    code: str,
    *,
    viewer_telegram_id: int,
    is_admin_view: bool = False,
) -> dict[str, Any]:
    """Xona holatini qaytaradi.

    Ishtirokchi uchun: joriy savol matni + deadline (javob yashirin).
    Admin uchun: barcha javoblar + to'g'ri javob ham ko'rinadi.
    Leak himoyasi: aktiv savol bo'lsa boshqa ishtirokchilarning javobi yashiriladi.
    """
    room = (
        GameRoom.objects.select_related("current_question")
        .filter(code=code.strip().upper())
        .first()
    )
    if not room:
        raise AppError(404, "Xona topilmadi")

    viewer_is_admin = room.is_room_admin(viewer_telegram_id)
    effective_admin_view = is_admin_view and viewer_is_admin

    participants = list(
        room.participants.all().order_by("-total_points", "speed_score_ms", "joined_at")
    )

    current_q = room.current_question
    current_q_data = None
    if current_q:
        current_q_data = _serialize_question_for_participant(
            current_q,
            viewer_telegram_id=viewer_telegram_id,
            is_admin_view=effective_admin_view,
            now_ms=_now_ms(),
        )

    return {
        "code": room.code,
        "name": room.name,
        "status": room.status,
        "adminTelegramId": room.admin_telegram_id,
        "hasPassword": bool(room.join_password),
        "createdAt": room.created_at.isoformat(),
        "startedAt": room.started_at.isoformat() if room.started_at else None,
        "finishedAt": room.finished_at.isoformat() if room.finished_at else None,
        "participantCount": len(participants),
        "leaderboard": [_serialize_participant(p, rank=i + 1) for i, p in enumerate(participants)],
        "currentQuestion": current_q_data,
        "viewerTelegramId": viewer_telegram_id,
        "viewerIsAdmin": viewer_is_admin,
    }


# ─── Guruhlangan baholash (admin uchun) ──────────────────────────────────────

def _normalize_answer(text: str) -> str:
    """Javob matnini normalizatsiya qiladi (grading.py bilan muvofiq).

    NFC + trim + ichki bo'shliqlarni yig'ish + casefold.
    casefold() lower() ga qaraganda ko'proq Unicode holatlarni to'g'ri ishlaydi
    (masalan, nemischa ß → ss), O'zbek/Kirilcha uchun ham xavfsiz.
    """
    import unicodedata
    import re
    if not text:
        return ""
    normalized = unicodedata.normalize("NFC", text).strip().casefold()
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def get_grouped_submissions(
    *,
    code: str,
    admin_telegram_id: int,
    question_id: int,
) -> dict[str, Any]:
    """Admin uchun savolning javoblarini guruhlaydi.

    Bir xil (normalized) javoblar bitta guruhga birlashtiriladi.
    Guruhlar count bo'yicha kamayish tartibida qaytariladi.

    Response shape:
    {
        questionId, questionBody, correctAnswer, questionStatus,
        totalSubmissions, totalUngraded,
        groups: [
            {
                normalizedKey: str,
                displayAnswer: str,       # original cased bir namunaviy javob
                count: int,
                gradedCorrect: int,
                gradedIncorrect: int,
                pending: int,
                submissionIds: list[int],
            },
            ...
        ]
    }
    """
    normalized_code = code.strip().upper()
    room = get_room(normalized_code)
    _require_room_admin(room, admin_telegram_id)

    question = GameQuestion.objects.filter(id=question_id, room=room).first()
    if not question:
        raise AppError(404, "Savol topilmadi")

    submissions = list(
        Submission.objects.select_related("participant")
        .filter(question=question)
        .order_by("submitted_at")
    )

    # Javoblarni normalize qilib guruhlash
    # groups_map: normalizedKey → {displayAnswer, submissionIds, counts}
    groups_map: dict[str, dict[str, Any]] = {}
    for sub in submissions:
        key = _normalize_answer(sub.answer_text)
        if key not in groups_map:
            groups_map[key] = {
                "normalizedKey": key,
                "displayAnswer": sub.answer_text,  # birinchi ko'ringan asl javob
                "submissionIds": [],
                "gradedCorrect": 0,
                "gradedIncorrect": 0,
                "pending": 0,
            }
        g = groups_map[key]
        g["submissionIds"].append(sub.id)
        if sub.is_correct is True:
            g["gradedCorrect"] += 1
        elif sub.is_correct is False:
            g["gradedIncorrect"] += 1
        else:
            g["pending"] += 1

    # Count bo'yicha kamayish tartibida saralash
    groups = sorted(groups_map.values(), key=lambda g: len(g["submissionIds"]), reverse=True)
    for g in groups:
        g["count"] = len(g["submissionIds"])

    total = len(submissions)
    total_ungraded = sum(1 for s in submissions if not s.graded_by)

    return {
        "questionId": question_id,
        "questionBody": question.body,
        "correctAnswer": question.correct_answer,
        "questionStatus": question.status,
        "totalSubmissions": total,
        "totalUngraded": total_ungraded,
        "groups": groups,
    }


def bulk_grade_group(
    *,
    code: str,
    admin_telegram_id: int,
    question_id: int,
    normalized_key: str,
    is_correct: bool,
) -> dict[str, Any]:
    """Bir normalized javob guruhining barcha submissionlarini baholaydi.

    Atomik va race-safe. Re-grading to'g'ri delta bilan ishlaydi:
    - Allaqachon to'g'ri deb belgilangan + yana to'g'ri → delta 0 (points o'zgarmaydi)
    - Allaqachon noto'g'ri + endi to'g'ri → delta = +point_value
    - Allaqachon to'g'ri + endi noto'g'ri → delta = -point_value, speed_score ham ayiriladi
    - Hali baholanmagan → oddiy grade

    Returns: { gradedCount, questionId, isCorrect }
    """
    normalized_code = code.strip().upper()
    room = get_room(normalized_code)
    _require_room_admin(room, admin_telegram_id)

    question = GameQuestion.objects.filter(id=question_id, room=room).first()
    if not question:
        raise AppError(404, "Savol topilmadi")

    # Ushbu question uchun normalized key mos keladigan barcha submissionlar
    # Normalizatsiyani Python tomonida bajaramiz (DB'da casefold/NFC yo'q)
    all_subs = list(
        Submission.objects.select_related("participant")
        .filter(question=question)
    )

    target_ids = [
        sub.id for sub in all_subs
        if _normalize_answer(sub.answer_text) == normalized_key
    ]

    if not target_ids:
        return {"gradedCount": 0, "questionId": question_id, "isCorrect": is_correct}

    new_points = question.point_value if is_correct else 0
    graded_count = 0

    with transaction.atomic():
        subs_to_grade = list(
            Submission.objects.select_for_update()
            .select_related("participant")
            .filter(id__in=target_ids)
        )

        for sub in subs_to_grade:
            old_points = sub.points_awarded or 0
            old_is_correct = sub.is_correct  # None | True | False

            # Delta hisoblash
            delta = new_points - old_points

            # Submission yangilash
            sub.is_correct = is_correct
            sub.points_awarded = new_points
            sub.graded_by = "manual"
            sub.grading_note = ""
            sub.save(update_fields=[
                "is_correct", "points_awarded", "graded_by",
                "grading_note", "updated_at",
            ])
            graded_count += 1

            if delta != 0:
                Participant.objects.filter(id=sub.participant_id).update(
                    total_points=models_F("total_points") + delta
                )

            # Speed score: faqat yangi to'g'ri javob uchun qo'shiladi
            # (oldin noto'g'ri/baholanmagan edi, endi to'g'ri bo'ldi)
            if is_correct and old_is_correct is not True and question.activated_at:
                activated_ms = int(question.activated_at.timestamp() * 1000)
                submitted_ms = int(sub.submitted_at.timestamp() * 1000)
                response_ms = max(0, submitted_ms - activated_ms)
                Participant.objects.filter(id=sub.participant_id).update(
                    speed_score_ms=models_F("speed_score_ms") + response_ms
                )

            # Speed score: to'g'ri edi, endi noto'g'ri → teskari qaytarish
            if not is_correct and old_is_correct is True and question.activated_at:
                activated_ms = int(question.activated_at.timestamp() * 1000)
                submitted_ms = int(sub.submitted_at.timestamp() * 1000)
                response_ms = max(0, submitted_ms - activated_ms)
                Participant.objects.filter(id=sub.participant_id).update(
                    speed_score_ms=models_F("speed_score_ms") - response_ms
                )

        # Manfiy bo'lib qolmasligini kafolatlash
        Participant.objects.filter(
            id__in=[s.participant_id for s in subs_to_grade],
            total_points__lt=0
        ).update(total_points=0)
        Participant.objects.filter(
            id__in=[s.participant_id for s in subs_to_grade],
            speed_score_ms__lt=0
        ).update(speed_score_ms=0)

    return {"gradedCount": graded_count, "questionId": question_id, "isCorrect": is_correct}


def grade_rest_wrong(
    *,
    code: str,
    admin_telegram_id: int,
    question_id: int,
) -> dict[str, Any]:
    """Hali baholanmagan (graded_by='') barcha submissionlarni noto'g'ri deb belgilaydi.

    0 ball beriladi, speed_score o'zgarmaydi (noto'g'ri javob uchun speed qo'shilmagan).
    Atomik.

    Returns: { gradedCount, questionId }
    """
    normalized_code = code.strip().upper()
    room = get_room(normalized_code)
    _require_room_admin(room, admin_telegram_id)

    question = GameQuestion.objects.filter(id=question_id, room=room).first()
    if not question:
        raise AppError(404, "Savol topilmadi")

    with transaction.atomic():
        updated_count = Submission.objects.filter(
            question=question,
            graded_by="",
        ).update(
            is_correct=False,
            points_awarded=0,
            graded_by="manual",
            grading_note="grade_rest_wrong",
        )

    return {"gradedCount": updated_count, "questionId": question_id}


# ─── Javoblar oqimi (admin uchun) ─────────────────────────────────────────────

def get_question_submissions(
    *,
    code: str,
    admin_telegram_id: int,
    question_id: int,
) -> dict[str, Any]:
    """Admin aktiv/yopiq savolning barcha javoblarini ko'radi."""
    normalized = code.strip().upper()
    room = get_room(normalized)
    _require_room_admin(room, admin_telegram_id)

    question = GameQuestion.objects.filter(id=question_id, room=room).first()
    if not question:
        raise AppError(404, "Savol topilmadi")

    submissions = list(
        Submission.objects.select_related("participant")
        .filter(question=question)
        .order_by("submitted_at")
    )

    return {
        "questionId": question_id,
        "questionBody": question.body,
        "correctAnswer": question.correct_answer,
        "status": question.status,
        "submissions": [_serialize_submission_for_admin(s) for s in submissions],
    }


# ─── Leaderboard ─────────────────────────────────────────────────────────────

def get_leaderboard(code: str, *, viewer_telegram_id: int) -> dict[str, Any]:
    """Joriy reyting — barcha ishtirokchilar ballar bo'yicha saralangan."""
    room = get_room(code)
    participants = list(
        room.participants.all().order_by("-total_points", "speed_score_ms", "joined_at")
    )
    winners = []
    if room.status == "finished":
        winners = [
            _serialize_participant(p, rank=i + 1)
            for i, p in enumerate(participants[:TOP_WINNERS_COUNT])
        ]
    return {
        "roomCode": room.code,
        "roomName": room.name,
        "status": room.status,
        "leaderboard": [_serialize_participant(p, rank=i + 1) for i, p in enumerate(participants)],
        "winners": winners,
        "viewerTelegramId": viewer_telegram_id,
    }


# ─── Statistika ──────────────────────────────────────────────────────────────

def get_room_stats(code: str, *, admin_telegram_id: int) -> dict[str, Any]:
    """Xona statistikasi — admin uchun."""
    normalized = code.strip().upper()
    room = get_room(normalized)
    _require_room_admin(room, admin_telegram_id)

    questions = list(room.questions.all())
    participants = list(room.participants.all())

    q_stats = []
    for q in questions:
        subs = list(q.submissions.all())
        total = len(subs)
        correct = sum(1 for s in subs if s.is_correct is True)
        q_stats.append({
            "questionId": q.id,
            "body": q.body[:80],
            "orderIndex": q.order_index,
            "totalSubmissions": total,
            "correctCount": correct,
            "incorrectCount": sum(1 for s in subs if s.is_correct is False),
            "pendingCount": sum(1 for s in subs if s.is_correct is None),
            "correctRate": round(correct / total * 100, 1) if total else 0,
        })

    # Eng to'g'ri javoblar olingan savol
    most_correct = max(q_stats, key=lambda x: x["correctCount"], default=None)
    # Eng qiyin savol (eng kam to'g'ri, kamida 1 submission)
    hardest = min(
        (x for x in q_stats if x["totalSubmissions"] > 0),
        key=lambda x: x["correctRate"],
        default=None,
    )

    # Har o'yinchi statistikasi
    player_stats = []
    for p in participants:
        p_subs = list(p.submissions.select_related("question").all())
        p_correct = sum(1 for s in p_subs if s.is_correct is True)
        player_stats.append({
            "telegramId": p.telegram_id,
            "displayName": p.display_name,
            "totalPoints": p.total_points,
            "totalSubmissions": len(p_subs),
            "correctCount": p_correct,
            "accuracy": round(p_correct / len(p_subs) * 100, 1) if p_subs else 0,
        })

    return {
        "roomCode": room.code,
        "roomName": room.name,
        "status": room.status,
        "participantCount": len(participants),
        "questionCount": len(questions),
        "questionStats": q_stats,
        "mostCorrectQuestion": most_correct,
        "hardestQuestion": hardest,
        "playerStats": player_stats,
    }


# ─── Eksport (Excel/CSV uchun raw data) ──────────────────────────────────────

def get_room_results(code: str, *, admin_telegram_id: int) -> dict[str, Any]:
    """Xona natijalarini eksport qilish uchun raw data.

    Excel/CSV render qiluvchi frontend/bot ushbu endpointni ishlatadi.
    """
    normalized = code.strip().upper()
    room = get_room(normalized)
    _require_room_admin(room, admin_telegram_id)

    questions = list(room.questions.prefetch_related("submissions__participant").all())
    participants = list(
        room.participants.all().order_by("-total_points", "speed_score_ms", "joined_at")
    )

    # Matrix: [participant][question] = submission data
    rows = []
    for rank, p in enumerate(participants, 1):
        row = {
            "rank": rank,
            "telegramId": p.telegram_id,
            "displayName": p.display_name,
            "totalPoints": p.total_points,
            "answers": [],
        }
        for q in questions:
            sub = next(
                (s for s in q.submissions.all() if s.participant_id == p.id),
                None,
            )
            row["answers"].append({
                "questionId": q.id,
                "orderIndex": q.order_index,
                "answerText": sub.answer_text if sub else None,
                "isCorrect": sub.is_correct if sub else None,
                "pointsAwarded": sub.points_awarded if sub else None,
                "gradedBy": sub.graded_by if sub else None,
                "submittedAt": sub.submitted_at.isoformat() if sub else None,
            })
        rows.append(row)

    return {
        "roomCode": room.code,
        "roomName": room.name,
        "status": room.status,
        "startedAt": room.started_at.isoformat() if room.started_at else None,
        "finishedAt": room.finished_at.isoformat() if room.finished_at else None,
        "questions": [
            {
                "id": q.id,
                "orderIndex": q.order_index,
                "body": q.body,
                "correctAnswer": q.correct_answer,
                "pointValue": q.point_value,
                "timeLimitSeconds": q.time_limit_seconds,
                "isBonus": q.is_bonus,
            }
            for q in questions
        ],
        "participants": rows,
    }


# ─── Yordamchi (private) funksiyalar ─────────────────────────────────────────

def _require_room_admin(room: GameRoom, telegram_id: int) -> None:
    if not room.is_room_admin(telegram_id):
        raise AppError(403, "Faqat xona admini bu amalni bajarishi mumkin")


def _add_participant_points(
    *,
    participant_id: int,
    points: int,
    submitted_at_ms: int,
    activated_at_ms: int,
) -> None:
    """Ishtirokchi balini va tezlik scoreni atomik yangilaydi."""
    response_ms = max(0, submitted_at_ms - activated_at_ms)
    Participant.objects.filter(id=participant_id).update(
        total_points=models_F("total_points") + points,
        speed_score_ms=models_F("speed_score_ms") + response_ms,
    )


def _deadline_info(question: GameQuestion, now_ms: int) -> tuple[int, bool]:
    """(time_remaining_ms, is_expired) qaytaradi."""
    if not question.activated_at:
        return 0, False
    deadline_ms = (
        int(question.activated_at.timestamp() * 1000)
        + question.time_limit_seconds * 1000
    )
    remaining = deadline_ms - now_ms
    return max(0, remaining), remaining < -GRACE_MS


# ─── Serializatsiya ──────────────────────────────────────────────────────────

def _serialize_participant(p: Participant, *, rank: int) -> dict[str, Any]:
    return {
        "rank": rank,
        "telegramId": p.telegram_id,
        "displayName": p.display_name,
        "totalPoints": p.total_points,
        "joinedAt": p.joined_at.isoformat(),
    }


def _serialize_question_for_admin(q: GameQuestion) -> dict[str, Any]:
    """Admin uchun: to'g'ri javob + barcha maydonlar."""
    return {
        "id": q.id,
        "roomCode": q.room.code if hasattr(q, "room") else None,
        "questionType": q.question_type,
        "body": q.body,
        "mediaRef": q.media_ref,
        "caption": q.caption,
        "correctAnswer": q.correct_answer,
        "timeLimitSeconds": q.time_limit_seconds,
        "pointValue": q.point_value,
        "orderIndex": q.order_index,
        "isBonus": q.is_bonus,
        "isQuick": q.is_quick,
        "status": q.status,
        "activatedAt": q.activated_at.isoformat() if q.activated_at else None,
        "closedAt": q.closed_at.isoformat() if q.closed_at else None,
        "createdAt": q.created_at.isoformat(),
    }


def _serialize_question_for_participant(
    q: GameQuestion,
    *,
    viewer_telegram_id: int,
    is_admin_view: bool,
    now_ms: int,
) -> dict[str, Any] | None:
    """Ishtirokchi uchun: to'g'ri javob YASHIRIN (savol yopilgunga qadar).

    Leak himoyasi:
    - `correctAnswer` faqat status='closed' bo'lsa yoki admin ko'rsa ochiladi.
    - Boshqa ishtirokchilarning `answerText` faqat savol yopilgandan keyin ko'rinadi.
    - `mediaUrl` — media_ref Telegram file_id bo'lsa proxy URL, http(s) bo'lsa to'g'ridan-to'g'ri.
    """
    if q is None:
        return None

    time_remaining_ms, expired = _deadline_info(q, now_ms)
    is_open = q.status == "active" and not expired

    # Ishtirokchi o'z javobini topish
    my_submission = None
    if viewer_telegram_id > 0:
        try:
            # Lazy select — faqat kerak bo'lganda
            participant = Participant.objects.filter(
                room_id=q.room_id, telegram_id=viewer_telegram_id
            ).first()
            if participant:
                sub = Submission.objects.filter(
                    question=q, participant=participant
                ).first()
                if sub:
                    my_submission = {
                        "submissionId": sub.id,
                        "answerText": sub.answer_text,
                        "submittedAt": sub.submitted_at.isoformat(),
                        "updatedAt": sub.updated_at.isoformat(),
                        "isCorrect": sub.is_correct if not is_open else None,
                        "pointsAwarded": sub.points_awarded if not is_open else None,
                    }
        except Exception:
            pass

    # To'g'ri javob: faqat admin yoki savol yopilganda
    reveal_correct = not is_open or is_admin_view

    # Media URL — Telegram file_id bo'lsa proxy, http(s) bo'lsa to'g'ridan
    media_url: str | None = None
    if q.media_ref:
        ref = q.media_ref.strip()
        if ref.startswith("http://") or ref.startswith("https://"):
            media_url = ref
        else:
            # Telegram file_id — backend proxy orqali
            room_code = q.room.code if hasattr(q, "room") and q.room else None
            if room_code:
                media_url = f"/api/gamerooms/rooms/{room_code}/questions/{q.id}/media"

    return {
        "id": q.id,
        "questionType": q.question_type,
        "body": q.body,
        "mediaRef": q.media_ref if q.media_ref else None,
        "mediaUrl": media_url,
        "caption": q.caption if q.caption else None,
        "timeLimitSeconds": q.time_limit_seconds,
        "pointValue": q.point_value,
        "orderIndex": q.order_index,
        "isBonus": q.is_bonus,
        "isQuick": q.is_quick,
        "status": q.status,
        "activatedAt": q.activated_at.isoformat() if q.activated_at else None,
        "closedAt": q.closed_at.isoformat() if q.closed_at else None,
        "timeRemainingMs": time_remaining_ms,
        "isExpired": expired,
        # Leak himoyasi — faqat yopilganda yoki admin ko'rsa
        "correctAnswer": q.correct_answer if reveal_correct and q.correct_answer else None,
        "mySubmission": my_submission,
    }


def _serialize_submission_for_admin(s: Submission) -> dict[str, Any]:
    return {
        "submissionId": s.id,
        "participantTelegramId": s.participant.telegram_id,
        "participantName": s.participant.display_name,
        "answerText": s.answer_text,
        "isCorrect": s.is_correct,
        "pointsAwarded": s.points_awarded,
        "gradedBy": s.graded_by,
        "gradingNote": s.grading_note,
        "submittedAt": s.submitted_at.isoformat(),
        "updatedAt": s.updated_at.isoformat(),
    }


# ─── 1. Excel eksport ─────────────────────────────────────────────────────────

def export_room_results_xlsx(code: str, *, admin_telegram_id: int):
    """Xona natijalarini haqiqiy .xlsx fayl sifatida qaytaradi.

    `openpyxl` ishlatiladi. Ikkita varaq:
      - Leaderboard: reyting, ism, ball, to'g'ri count, aniqlik %
      - Savollar: tartib, matn, tur, to'g'ri javob, statistika
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise AppError(500, "openpyxl kutubxonasi o'rnatilmagan — server admini bilan bog'laning")

    normalized = code.strip().upper()
    room = get_room(normalized)
    _require_room_admin(room, admin_telegram_id)

    questions = list(room.questions.prefetch_related("submissions__participant").all())
    participants = list(
        room.participants.all().order_by("-total_points", "speed_score_ms", "joined_at")
    )

    wb = openpyxl.Workbook()

    # ── Varaq 1: Leaderboard ──────────────────────────────────────────────────
    ws_lb = wb.active
    ws_lb.title = "Leaderboard"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center = Alignment(horizontal="center")

    lb_headers = ["#", "Ism", "Ball", "To'g'ri javoblar", "Aniqlik (%)"]
    for col, h in enumerate(lb_headers, 1):
        cell = ws_lb.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws_lb.column_dimensions["A"].width = 6
    ws_lb.column_dimensions["B"].width = 28
    ws_lb.column_dimensions["C"].width = 10
    ws_lb.column_dimensions["D"].width = 18
    ws_lb.column_dimensions["E"].width = 16

    for rank, p in enumerate(participants, 1):
        p_subs = list(p.submissions.all())
        correct_count = sum(1 for s in p_subs if s.is_correct is True)
        total_subs = len(p_subs)
        accuracy = round(correct_count / total_subs * 100, 1) if total_subs else 0.0
        ws_lb.append([rank, p.display_name, p.total_points, correct_count, accuracy])

    # ── Varaq 2: Savollar statistikasi ────────────────────────────────────────
    ws_q = wb.create_sheet("Savollar")

    q_headers = [
        "#", "Savol matni", "Tur", "To'g'ri javob",
        "Jami javob", "To'g'ri", "Noto'g'ri", "Baholanmagan", "To'g'ri %",
    ]
    for col, h in enumerate(q_headers, 1):
        cell = ws_q.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws_q.column_dimensions["A"].width = 6
    ws_q.column_dimensions["B"].width = 50
    ws_q.column_dimensions["C"].width = 10
    ws_q.column_dimensions["D"].width = 30
    ws_q.column_dimensions["E"].width = 12
    ws_q.column_dimensions["F"].width = 12
    ws_q.column_dimensions["G"].width = 14
    ws_q.column_dimensions["H"].width = 16
    ws_q.column_dimensions["I"].width = 12

    for q in questions:
        subs = list(q.submissions.all())
        total = len(subs)
        correct = sum(1 for s in subs if s.is_correct is True)
        incorrect = sum(1 for s in subs if s.is_correct is False)
        pending = sum(1 for s in subs if s.is_correct is None)
        rate = round(correct / total * 100, 1) if total else 0.0
        ws_q.append([
            q.order_index,
            q.body,
            q.question_type,
            q.correct_answer or "",
            total,
            correct,
            incorrect,
            pending,
            rate,
        ])

    # Wrap text uchun savollar varaqda
    for row in ws_q.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"gameroom_{normalized}_results.xlsx"


# ─── 2. Admin xonalari ro'yxati ──────────────────────────────────────────────

def list_admin_rooms(*, admin_telegram_id: int) -> list[dict[str, Any]]:
    """Admin o'zi yaratgan yoki co-admin bo'lgan barcha xonalarni qaytaradi.

    Filtrlash DB darajasida amalga oshirilmaydi (extra_admin_ids JSON maydon),
    shuning uchun owner bo'lgan xonalar DB'da filter qilinadi, co-admin bo'lganlar
    Python darajasida tekshiriladi.
    Ish hajmi amalda kichik (bir admin yuzlab xona yaratmaydi).
    """
    # Owner bo'lgan xonalar — DB filter
    owned = list(
        GameRoom.objects.filter(admin_telegram_id=admin_telegram_id)
        .order_by("-created_at")
    )

    # Co-admin bo'lgan xonalar — barcha non-owned xonalarni tekshirib chiqish
    # muammo ko'rsatishi mumkin, lekin amalda xonalar soni cheklangan.
    # Optimizatsiya kerak bo'lsa keyinchalik DB indexi qo'shish mumkin.
    co_admin_rooms = list(
        GameRoom.objects.exclude(admin_telegram_id=admin_telegram_id)
        .order_by("-created_at")
    )
    co_admin_filtered = [
        r for r in co_admin_rooms
        if r.is_room_admin(admin_telegram_id)
    ]

    # Barcha xonalar — owned birinchi, co-admin ikkinchi, created_at bo'yicha
    all_rooms = owned + co_admin_filtered
    all_rooms.sort(key=lambda r: r.created_at, reverse=True)

    result = []
    for r in all_rooms:
        participant_count = r.participants.count()
        question_count = r.questions.count()
        result.append({
            "code": r.code,
            "name": r.name,
            "status": r.status,
            "isOwner": r.admin_telegram_id == admin_telegram_id,
            "hasPassword": bool(r.join_password),
            "participantCount": participant_count,
            "questionCount": question_count,
            "createdAt": r.created_at.isoformat(),
            "startedAt": r.started_at.isoformat() if r.started_at else None,
            "finishedAt": r.finished_at.isoformat() if r.finished_at else None,
        })
    return result


# ─── 3. Media proxy ──────────────────────────────────────────────────────────

def resolve_question_media(
    *,
    code: str,
    question_id: int,
    viewer_telegram_id: int,
) -> tuple[bytes, str]:
    """Savol media faylini Telegram'dan stream qilib qaytaradi.

    XAVFSIZLIK:
    - Faqat ushbu xonaga tegishli savol media_ref'i proxylash mumkin
      (ixtiyoriy file_id qabul qilinmaydi — ochiq proxy yo'q).
    - Foydalanuvchi xonada ro'yxatdan o'tgan bo'lishi shart.

    Returns: (bytes_content, content_type)
    """
    import urllib.request
    from django.conf import settings

    normalized = code.strip().upper()

    # Savol topilishi shart
    room = get_room(normalized)
    question = GameQuestion.objects.filter(id=question_id, room=room).first()
    if not question:
        raise AppError(404, "Savol topilmadi")

    media_ref = (question.media_ref or "").strip()
    if not media_ref:
        raise AppError(404, "Bu savolda media yo'q")

    # Foydalanuvchi xonada ro'yxatdan o'tganligini tekshirish — AVVAL auth, keyin media
    # (admin ham ishtirokchi sifatida kirgan bo'lishi shart emas — admin view uchun ruxsat)
    viewer_is_admin = room.is_room_admin(viewer_telegram_id)
    if not viewer_is_admin:
        participant = Participant.objects.filter(
            room=room, telegram_id=viewer_telegram_id
        ).first()
        if not participant:
            raise AppError(403, "Siz bu xonada ro'yxatdan o'tmagansiz")

    # Agar allaqachon http URL bo'lsa — tekshiruvdan O'TGANDAN SO'NG redirect
    if media_ref.startswith("http://") or media_ref.startswith("https://"):
        raise AppError(302, media_ref)  # view tomonida redirect qilinadi

    # Bot token tekshirish
    bot_token = getattr(settings, "TELEGRAM_BOT_TOKEN", "")
    if not bot_token:
        raise AppError(502, "Bot token sozlanmagan — server admini bilan bog'laning")

    # Telegram getFile API — file_path olish
    get_file_url = f"https://api.telegram.org/bot{bot_token}/getFile?file_id={media_ref}"
    try:
        with urllib.request.urlopen(get_file_url, timeout=10) as resp:  # noqa: S310
            import json as _json
            data = _json.loads(resp.read().decode())
    except Exception as e:
        logger.warning("Telegram getFile xatosi: %s", e)
        raise AppError(502, "Telegram media manzilini olishda xato")

    if not data.get("ok"):
        raise AppError(404, "Telegram file topilmadi (file_id eskirgan yoki noto'g'ri)")

    file_path = data.get("result", {}).get("file_path", "")
    if not file_path:
        raise AppError(502, "Telegram file_path bo'sh qaytdi")

    # Fayl yuklab olish — hajm va timeout cheklovi bilan
    _MAX_MEDIA_BYTES = 20 * 1024 * 1024  # 20 MB (Telegram bot limit)
    download_url = f"https://api.telegram.org/file/bot{bot_token}/{file_path}"
    try:
        with urllib.request.urlopen(download_url, timeout=10) as resp:  # noqa: S310
            content_length = resp.headers.get("Content-Length")
            if content_length is not None and int(content_length) > _MAX_MEDIA_BYTES:
                raise AppError(413, "Media fayli juda katta (20 MB dan oshmasligi kerak)")
            # Content-Length yo'q yoki ishonchsiz bo'lsa ham cap bilan o'qiymiz
            content = resp.read(_MAX_MEDIA_BYTES + 1)
            if len(content) > _MAX_MEDIA_BYTES:
                raise AppError(413, "Media fayli juda katta (20 MB dan oshmasligi kerak)")
            content_type = resp.headers.get("Content-Type", "application/octet-stream")
    except AppError:
        raise
    except Exception as e:
        logger.warning("Telegram fayl yuklab olishda xato: %s", e)
        raise AppError(502, "Telegram faylni yuklab olishda xato")

    return content, content_type


# ─── 4. Savollarni savol bankiga saqlash ────────────────────────────────────

def save_questions_to_bank(
    *,
    code: str,
    admin_telegram_id: int,
    category: str | None = None,
    difficulty: str | None = None,
) -> dict[str, Any]:
    """Xonaning matn savollarini apps.questions bankiga saqlaydi.

    Qoidalar:
    - Faqat question_type='text' savollar saqlanadi.
    - correct_answer bo'sh bo'lgan savollar o'tkaziladi.
    - Allaqachon bank_question_id to'ldirilgan savollar o'tkaziladi (idempotent).
    - Aniq dublikat oldini olish: bir xil (text, correct_answer) kombinatsiyasi
      bankda mavjud bo'lsa, GameQuestion.bank_question_id yangilanadi, lekin
      yangi yozuv yaratilmaydi.

    Returns: { savedCount, skippedCount, alreadySavedCount }
    """
    from apps.questions.models import Question

    normalized = code.strip().upper()
    room = get_room(normalized)
    _require_room_admin(room, admin_telegram_id)

    questions = list(
        GameQuestion.objects.filter(
            room=room,
            question_type="text",
        ).exclude(correct_answer="")
    )

    saved_count = 0
    skipped_count = 0
    already_saved_count = 0

    for gq in questions:
        # Allaqachon saqlangan
        if gq.bank_question_id is not None:
            already_saved_count += 1
            continue

        body_stripped = gq.body.strip()
        answer_stripped = gq.correct_answer.strip()

        # Dublikat tekshirish — bo'shliq va katta-kichik harf farqini e'tiborga olmasdan
        existing_bank_q = Question.objects.filter(
            text__iexact=body_stripped,
            correct_answer__iexact=answer_stripped,
        ).first()

        with transaction.atomic():
            if existing_bank_q:
                # Bankda mavjud — faqat FK belgilaymiz
                GameQuestion.objects.filter(id=gq.id).update(
                    bank_question_id=existing_bank_q.id
                )
                skipped_count += 1
            else:
                # Yangi yozuv yaratish
                new_q = Question.objects.create(
                    text=body_stripped,
                    correct_answer=answer_stripped,
                    category=category,
                    difficulty=difficulty,
                    wrong_answers=[],
                    time_limit_seconds=None,
                )
                GameQuestion.objects.filter(id=gq.id).update(
                    bank_question_id=new_q.id
                )
                saved_count += 1

    return {
        "savedCount": saved_count,
        "skippedCount": skipped_count,
        "alreadySavedCount": already_saved_count,
        "totalTextQuestions": len(questions),
    }


# ─── 5. Savolni bekor qilish (cancel) ───────────────────────────────────────

def cancel_question(
    *,
    code: str,
    admin_telegram_id: int,
    question_id: int,
) -> dict[str, Any]:
    """Savolni bekor qiladi — barcha submissionlar va ball ta'siri o'chiriladi.

    Nima qiladi:
    1. Savolni topadi (pending yoki active).
    2. Agar allaqachon baholangan submissionlar bo'lsa — ular bergan ballni
       ishtirokchi total_points'idan atomik ayiradi.
    3. speed_score_ms ni ham to'g'ri javob submission'larning hissasini ayiradi.
    4. Barcha submission'larni o'chiradi.
    5. Savolning o'zini o'chiradi.
    6. room.current_question shu savol bo'lsa — null qiladi.

    Atomik: bitta transaction ichida.
    Idempotent: savol topilmasa 404 qaytaradi.
    """
    normalized = code.strip().upper()

    with transaction.atomic():
        room = GameRoom.objects.select_for_update().filter(code=normalized).first()
        if not room:
            raise AppError(404, "Xona topilmadi")
        _require_room_admin(room, admin_telegram_id)

        question = GameQuestion.objects.select_for_update().filter(
            id=question_id, room=room
        ).first()
        if not question:
            raise AppError(404, "Savol topilmadi")
        if question.status == "closed":
            # Yopilgan savollarni ham bekor qilishga ruxsat beramiz (admin qaroriga)
            pass  # allowed

        # Baholangan submissionlarning ball ta'sirini teskari qaytarish
        graded_subs = list(
            Submission.objects.select_for_update()
            .select_related("participant")
            .filter(question=question, graded_by__in=["auto", "manual"])
        )

        for sub in graded_subs:
            awarded = sub.points_awarded or 0
            # Ball qaytarish
            if awarded > 0:
                Participant.objects.filter(id=sub.participant_id).update(
                    total_points=models_F("total_points") - awarded,
                )
            # Speed score qaytarish — faqat to'g'ri javoblar uchun qo'shilgandi
            if sub.is_correct and question.activated_at:
                activated_ms = int(question.activated_at.timestamp() * 1000)
                submitted_ms = int(sub.submitted_at.timestamp() * 1000)
                response_ms = max(0, submitted_ms - activated_ms)
                Participant.objects.filter(id=sub.participant_id).update(
                    speed_score_ms=models_F("speed_score_ms") - response_ms,
                )

        # total_points manfiy bo'lib qolmasligini kafolatlash (edge case)
        Participant.objects.filter(room=room, total_points__lt=0).update(total_points=0)
        Participant.objects.filter(room=room, speed_score_ms__lt=0).update(speed_score_ms=0)

        # Barcha submissionlarni o'chirish
        deleted_subs_count = Submission.objects.filter(question=question).delete()[0]

        # room.current_question ni tozalash
        if room.current_question_id == question.id:
            room.current_question = None
            room.save(update_fields=["current_question"])

        # Savolni o'chirish
        q_order = question.order_index
        q_body_preview = question.body[:60]
        question.delete()

    return {
        "cancelledQuestionId": question_id,
        "orderIndex": q_order,
        "bodyPreview": q_body_preview,
        "deletedSubmissionsCount": deleted_subs_count,
        "pointsReversedFor": len([s for s in graded_subs if (s.points_awarded or 0) > 0]),
    }

